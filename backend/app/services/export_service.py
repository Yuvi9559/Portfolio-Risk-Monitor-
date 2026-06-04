from __future__ import annotations

import io
import logging
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─────────────────────────────────────────────────────────────────────────────
# PDF Report
# ─────────────────────────────────────────────────────────────────────────────

_HEADER_COLOUR = colors.HexColor("#1a3c5e")
_ACCENT_COLOUR = colors.HexColor("#2563eb")
_ALT_ROW_COLOUR = colors.HexColor("#f0f4ff")


def _make_table_style(header_bg: colors.Color = _HEADER_COLOUR) -> TableStyle:
    return TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), header_bg),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, _ALT_ROW_COLOUR]),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#d1d5db")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
        ]
    )


def generate_pdf(
    portfolio_name: str,
    holdings_data: List[Dict[str, Any]],
    risk_data: Dict[str, Any],
) -> bytes:
    """Generate a PDF risk report and return raw bytes."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
        title=f"Risk Report – {portfolio_name}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=20,
        textColor=_HEADER_COLOUR,
        spaceAfter=6,
    )
    heading_style = ParagraphStyle(
        "CustomHeading",
        parent=styles["Heading2"],
        fontSize=13,
        textColor=_HEADER_COLOUR,
        spaceAfter=4,
        spaceBefore=10,
    )
    normal_style = styles["Normal"]

    story = []

    # ── Title ─────────────────────────────────────────────────────────────────
    story.append(Paragraph(f"Portfolio Risk Report", title_style))
    story.append(Paragraph(f"<b>{portfolio_name}</b>", styles["Heading1"]))
    story.append(
        Paragraph(
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
            normal_style,
        )
    )
    story.append(HRFlowable(width="100%", thickness=1, color=_ACCENT_COLOUR))
    story.append(Spacer(1, 12))

    # ── Portfolio Summary ─────────────────────────────────────────────────────
    if risk_data:
        story.append(Paragraph("Portfolio Summary", heading_style))
        pv = risk_data.get("portfolio_value", 0)
        summary_data = [
            ["Metric", "Value"],
            ["Portfolio Value", f"${pv:,.2f}"],
            ["Daily Return", f"{risk_data.get('daily_return', 0):.2f}%"],
            ["Annualised Volatility", f"{risk_data.get('volatility', 0):.2f}%"],
            ["Sharpe Ratio", f"{risk_data.get('sharpe', 0):.2f}"],
            ["Sortino Ratio", f"{risk_data.get('sortino', 0):.2f}"],
            ["Beta", f"{risk_data.get('beta', 0):.2f}"],
            ["Max Drawdown", f"{risk_data.get('max_drawdown', 0):.2f}%"],
            ["VaR 95% (1-day)", f"{risk_data.get('var_95', 0):.2f}%  /  ${risk_data.get('var_95_dollar', 0):,.2f}"],
            ["CVaR 95%", f"{risk_data.get('cvar_95', 0):.2f}%"],
            ["VaR 99% (1-day)", f"{risk_data.get('var_99', 0):.2f}%"],
        ]
        tbl = Table(summary_data, colWidths=[8 * cm, 8 * cm])
        tbl.setStyle(_make_table_style())
        story.append(tbl)
        story.append(Spacer(1, 12))

    # ── Holdings ──────────────────────────────────────────────────────────────
    if holdings_data:
        story.append(Paragraph("Holdings", heading_style))
        h_header = [
            "Symbol", "Type", "Shares", "Avg Cost", "Current Price",
            "Market Value", "P&L $", "P&L %",
        ]
        rows = [h_header]
        for h in holdings_data:
            rows.append(
                [
                    h.get("symbol", ""),
                    h.get("asset_type", ""),
                    f"{h.get('shares', 0):.4f}",
                    f"${h.get('avg_cost') or 0:.2f}" if h.get("avg_cost") else "—",
                    f"${h.get('current_price') or 0:.2f}" if h.get("current_price") else "—",
                    f"${h.get('market_value') or 0:,.2f}" if h.get("market_value") else "—",
                    f"${h.get('pnl_dollar') or 0:,.2f}" if h.get("pnl_dollar") is not None else "—",
                    f"{h.get('pnl_pct') or 0:.2f}%" if h.get("pnl_pct") is not None else "—",
                ]
            )
        tbl2 = Table(rows, colWidths=[2.2 * cm, 1.8 * cm, 2 * cm, 2.2 * cm, 2.5 * cm, 2.5 * cm, 2 * cm, 2 * cm])
        tbl2.setStyle(_make_table_style())
        story.append(tbl2)
        story.append(Spacer(1, 12))

    # ── Footer ────────────────────────────────────────────────────────────────
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.grey))
    story.append(
        Paragraph(
            "<i>This report is generated automatically by Portfolio Risk Monitor Pro. "
            "Past performance is not indicative of future results.</i>",
            ParagraphStyle("Footer", parent=normal_style, fontSize=7, textColor=colors.grey),
        )
    )

    doc.build(story)
    return buffer.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Excel Report
# ─────────────────────────────────────────────────────────────────────────────

def _header_style(ws, row: int, columns: List[str]) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="1a3c5e")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                cell_len = len(str(cell.value or ""))
                if cell_len > max_len:
                    max_len = cell_len
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def generate_excel(
    portfolio_name: str,
    holdings_data: List[Dict[str, Any]],
    risk_data: Dict[str, Any],
    history_data: List[Dict[str, Any]],
) -> bytes:
    """Generate an Excel risk report and return raw bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Holdings ─────────────────────────────────────────────────────
    ws_h = wb.active
    ws_h.title = "Holdings"
    ws_h["A1"] = f"Portfolio Risk Report – {portfolio_name}"
    ws_h["A1"].font = Font(bold=True, size=14, color="1a3c5e")
    ws_h["A2"] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws_h["A2"].font = Font(italic=True, size=10, color="555555")

    h_cols = ["Symbol", "Asset Type", "Shares", "Avg Cost", "Current Price", "Market Value", "P&L $", "P&L %"]
    _header_style(ws_h, 4, h_cols)

    alt_fill = PatternFill(fill_type="solid", fgColor="EFF6FF")
    for i, h in enumerate(holdings_data, start=5):
        row_data = [
            h.get("symbol", ""),
            h.get("asset_type", ""),
            h.get("shares", 0),
            h.get("avg_cost"),
            h.get("current_price"),
            h.get("market_value"),
            h.get("pnl_dollar"),
            h.get("pnl_pct"),
        ]
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_h.cell(row=i, column=col_idx, value=val)
            if i % 2 == 0:
                cell.fill = alt_fill
            cell.alignment = Alignment(horizontal="center")
    _auto_width(ws_h)

    # ── Sheet 2: Risk Metrics ─────────────────────────────────────────────────
    ws_r = wb.create_sheet("Risk Metrics")
    ws_r["A1"] = "Risk Metrics"
    ws_r["A1"].font = Font(bold=True, size=14, color="1a3c5e")

    metric_rows = [
        ("Portfolio Value", f"${risk_data.get('portfolio_value', 0):,.2f}"),
        ("Daily Return", f"{risk_data.get('daily_return', 0):.2f}%"),
        ("Annualised Volatility", f"{risk_data.get('volatility', 0):.2f}%"),
        ("Sharpe Ratio", f"{risk_data.get('sharpe', 0):.4f}"),
        ("Sortino Ratio", f"{risk_data.get('sortino', 0):.4f}"),
        ("Beta", f"{risk_data.get('beta', 0):.4f}"),
        ("Max Drawdown", f"{risk_data.get('max_drawdown', 0):.2f}%"),
        ("VaR 95% (1-day %)", f"{risk_data.get('var_95', 0):.2f}%"),
        ("VaR 95% (1-day $)", f"${risk_data.get('var_95_dollar', 0):,.2f}"),
        ("CVaR 95%", f"{risk_data.get('cvar_95', 0):.2f}%"),
        ("VaR 99% (1-day %)", f"{risk_data.get('var_99', 0):.2f}%"),
        ("Report Timestamp", risk_data.get("ts", "")),
    ]
    _header_style(ws_r, 3, ["Metric", "Value"])
    for i, (metric, value) in enumerate(metric_rows, start=4):
        ws_r.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws_r.cell(row=i, column=2, value=value)
        if i % 2 == 0:
            for c in [1, 2]:
                ws_r.cell(row=i, column=c).fill = alt_fill
    _auto_width(ws_r)

    # ── Sheet 3: Risk History ─────────────────────────────────────────────────
    ws_hist = wb.create_sheet("Risk History")
    ws_hist["A1"] = "Historical Risk Snapshots"
    ws_hist["A1"].font = Font(bold=True, size=14, color="1a3c5e")

    hist_cols = ["Timestamp", "Portfolio Value", "VaR 95%", "Sharpe", "Volatility", "Max Drawdown"]
    _header_style(ws_hist, 3, hist_cols)
    for i, snap in enumerate(history_data, start=4):
        row_vals = [
            snap.get("ts", ""),
            snap.get("portfolio_value", 0),
            snap.get("var_95", 0),
            snap.get("sharpe", 0),
            snap.get("volatility", 0),
            snap.get("max_drawdown", 0),
        ]
        for col_idx, val in enumerate(row_vals, start=1):
            cell = ws_hist.cell(row=i, column=col_idx, value=val)
            cell.alignment = Alignment(horizontal="center")
            if i % 2 == 0:
                cell.fill = alt_fill
    _auto_width(ws_hist)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
